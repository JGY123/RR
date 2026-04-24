#!/usr/bin/env python3
"""
Rebuild data/security_ref.json from the source Excel.

Input:  data/security_flags_source.xlsx (sheet "19. Stocks raw exposure")
Output: data/security_ref.json  ({ meta, by_sedol: { SEDOL: {n, ccy, country, industry, tkr, isin} } })

Run: python3 data/bake_security_ref.py
Deterministic — same input always produces the same output (byte-identical up to dict ordering).

Coverage expectation vs a typical production CSV: ~97% via direct SEDOL match.
"""
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl; install via: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HERE = Path(__file__).parent
XLSX = HERE / "security_flags_source.xlsx"
OUT = HERE / "security_ref.json"
SHEET = "19. Stocks raw exposure"

# Known-country set (authoritative — derived 2026-04-24 from the EM+IDM+IOP+ACWI+SCG+ISC universes)
COUNTRY_SET = {
    "Australia", "Austria", "Belgium", "Brazil", "Canada", "Chile", "China", "Colombia",
    "Czech Republic", "Denmark", "Egypt", "Finland", "France", "Germany", "Greece",
    "Hong Kong", "Hungary", "India", "Indonesia", "Ireland", "Israel", "Italy", "Japan",
    "Korea, Republic of", "Kuwait", "Luxembourg", "Malaysia", "Mexico", "Netherlands",
    "New Zealand", "Nigeria", "Norway", "Peru", "Philippines", "Poland", "Portugal",
    "Qatar", "Romania", "Saudi Arabia", "Singapore", "South Africa", "Spain", "Sweden",
    "Switzerland", "Taiwan", "Thailand", "Turkey", "Ukraine", "United Arab Emirates",
    "United Kingdom", "United States",
    "Domestic China",  # Special marker used alongside China — kept as a country-type flag
}


def classify_column(header):
    """Return 'ccy' | 'country' | 'industry' for a flag-column header."""
    h = str(header).strip() if header else ""
    if len(h) == 3 and h.isalpha() and h.isupper():
        return "ccy"
    if h in COUNTRY_SET:
        return "country"
    return "industry"


def main():
    if not XLSX.exists():
        print(f"Source Excel not found at {XLSX}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]

    # Row 9 = header
    header_row = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    headers = list(header_row)

    # Identifier columns (indexes are 0-based): 0=name, 1=SEDOL, 2=SEDOLCHK, 3=Ticker-Region, 4=CUSIP, 5=ISIN
    # Flag columns start at index 6
    col_kind = []
    for i, h in enumerate(headers):
        if h is None:
            col_kind.append(("ignore", None))
            continue
        if i < 6:
            col_kind.append(("id", h))
            continue
        col_kind.append((classify_column(h), h))

    lookup = {}
    stats = {"rows_seen": 0, "no_sedol": 0, "section_headers": 0, "ambiguous": 0}

    # Data rows 11..max (with max_row large enough to cover the 6 concatenated strategy blocks)
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        name = row[0]
        if name is None:
            continue
        if isinstance(name, str) and name.startswith("Redwood"):
            stats["section_headers"] += 1
            continue
        sedol = row[1]
        if not sedol:
            stats["no_sedol"] += 1
            continue
        stats["rows_seen"] += 1

        ccy = country = industry = None
        for j, v in enumerate(row):
            if j >= len(col_kind):
                break
            if v in (1, 1.0, "1"):
                kind, hdr = col_kind[j]
                if kind == "ccy" and not ccy:
                    ccy = hdr
                elif kind == "country" and not country:
                    country = hdr
                elif kind == "industry" and not industry:
                    industry = hdr

        entry = {"n": name, "ccy": ccy, "country": country, "industry": industry}
        ticker = row[3]
        if ticker and ticker != "--":
            entry["tkr"] = ticker
        isin = row[5]
        if isin:
            entry["isin"] = isin

        # Dedupe — same security may appear across strategy universes with identical classification
        sedol_key = str(sedol)
        if sedol_key in lookup:
            prev = lookup[sedol_key]
            if (prev.get("ccy") != entry.get("ccy") or
                    prev.get("country") != entry.get("country") or
                    prev.get("industry") != entry.get("industry")):
                stats["ambiguous"] += 1
        else:
            lookup[sedol_key] = entry

    out = {
        "meta": {
            "source": XLSX.name,
            "sheet": SHEET,
            "schema_version": "1.0",
            "extracted_script": "data/bake_security_ref.py",
            "strategies_included": "EM, IDM, IOP, ACWI, SCG-variant, ISC (6 of 7 — GSC not in source)",
            "row_stats": stats,
            "ccy_count": sum(1 for k, _ in col_kind if k == "ccy"),
            "country_count": sum(1 for k, _ in col_kind if k == "country"),
            "industry_count": sum(1 for k, _ in col_kind if k == "industry"),
        },
        "by_sedol": lookup,
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=1, sort_keys=True)

    print(f"Wrote {len(lookup)} unique securities to {OUT}")
    print(f"Stats: {stats}")
    if stats["ambiguous"]:
        print(f"WARNING: {stats['ambiguous']} SEDOLs had ambiguous classifications across strategies.")


if __name__ == "__main__":
    main()
